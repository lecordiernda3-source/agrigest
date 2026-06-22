"""
AgriGest - Application de gestion des dépenses agricoles par client
Lancement : python agrigest2.py  puis ouvrir http://localhost:5000
"""

import sqlite3, os, json
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import webbrowser, threading

DB_FILE = "agrigest.db"

def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS clients (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT NOT NULL, telephone TEXT, culture TEXT,
        date_creation TEXT DEFAULT (date('now')))""")
    c.execute("""CREATE TABLE IF NOT EXISTS transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER NOT NULL,
        type TEXT NOT NULL CHECK(type IN ('depense','remboursement')),
        montant REAL NOT NULL, description TEXT,
        date TEXT DEFAULT (date('now')),
        FOREIGN KEY (client_id) REFERENCES clients(id))""")
    conn.commit(); conn.close()

def get_clients():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("""SELECT cl.id, cl.nom, cl.telephone, cl.culture,
        COALESCE(SUM(CASE WHEN t.type='depense' THEN t.montant ELSE 0 END),0),
        COALESCE(SUM(CASE WHEN t.type='remboursement' THEN t.montant ELSE 0 END),0)
        FROM clients cl LEFT JOIN transactions t ON cl.id=t.client_id
        GROUP BY cl.id ORDER BY cl.nom""")
    rows = c.fetchall(); conn.close()
    return [{"id":r[0],"nom":r[1],"telephone":r[2] or "","culture":r[3] or "",
             "total_depenses":r[4],"total_remboursements":r[5],"credit_restant":r[4]-r[5]} for r in rows]

def add_client(nom, telephone, culture):
    conn = sqlite3.connect(DB_FILE)
    conn.cursor().execute("INSERT INTO clients (nom,telephone,culture) VALUES (?,?,?)",(nom,telephone,culture))
    conn.commit(); conn.close()

def delete_client(client_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("DELETE FROM transactions WHERE client_id=?",(client_id,))
    c.execute("DELETE FROM clients WHERE id=?",(client_id,))
    conn.commit(); conn.close()

def get_transactions(client_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("SELECT id,type,montant,description,date FROM transactions WHERE client_id=? ORDER BY date DESC,id DESC",(client_id,))
    rows = c.fetchall(); conn.close()
    return [{"id":r[0],"type":r[1],"montant":r[2],"description":r[3] or "","date":r[4]} for r in rows]

def add_transaction(client_id, type_tx, montant, description, date):
    conn = sqlite3.connect(DB_FILE)
    conn.cursor().execute("INSERT INTO transactions (client_id,type,montant,description,date) VALUES (?,?,?,?,?)",
        (client_id,type_tx,montant,description,date))
    conn.commit(); conn.close()

def delete_transaction(tx_id):
    conn = sqlite3.connect(DB_FILE)
    conn.cursor().execute("DELETE FROM transactions WHERE id=?",(tx_id,))
    conn.commit(); conn.close()

def get_client_by_id(client_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("SELECT id,nom,telephone,culture FROM clients WHERE id=?",(client_id,))
    r = c.fetchone(); conn.close()
    return {"id":r[0],"nom":r[1],"telephone":r[2] or "","culture":r[3] or ""} if r else None

def export_excel(client_id):
    client = get_client_by_id(client_id)
    if not client: return None
    transactions = get_transactions(client_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport"
    green = "2E7D32"; light_green = "E8F5E9"; white = "FFFFFF"; gray = "F5F5F5"; light_red = "FFEBEE"
    def bd():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s,right=s,top=s,bottom=s)
    ws.merge_cells("A1:E1")
    c = ws["A1"]; c.value = f"AgriGest - Rapport : {client['nom']}"
    c.font = Font(bold=True,size=14,color=white); c.fill = PatternFill("solid",fgColor=green)
    c.alignment = Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height = 32
    ws.merge_cells("A2:E2")
    ws["A2"].value = f"Culture: {client['culture']} | Tel: {client['telephone']} | Exporte le {datetime.now().strftime('%d/%m/%Y')}"
    ws["A2"].alignment = Alignment(horizontal="center"); ws["A2"].fill = PatternFill("solid",fgColor=light_green)
    headers = ["Date","Type","Description","Montant (FCFA)","Solde cumule"]
    for col,h in enumerate(headers,1):
        cell = ws.cell(row=4,column=col,value=h)
        cell.font = Font(bold=True,color=white); cell.fill = PatternFill("solid",fgColor="388E3C")
        cell.alignment = Alignment(horizontal="center"); cell.border = bd()
    solde = 0
    for i,tx in enumerate(reversed(transactions),5):
        if tx["type"]=="depense": solde+=tx["montant"]; fc=light_red; tl="Depense"
        else: solde-=tx["montant"]; fc=light_green; tl="Remboursement"
        fill = PatternFill("solid",fgColor=fc if i%2==0 else white)
        for col,val in enumerate([tx["date"],tl,tx["description"],tx["montant"],solde],1):
            cell = ws.cell(row=i,column=col,value=val)
            cell.fill = fill; cell.border = bd()
            cell.alignment = Alignment(horizontal="center" if col in [1,2,4,5] else "left")
    total_dep = sum(t["montant"] for t in transactions if t["type"]=="depense")
    total_remb = sum(t["montant"] for t in transactions if t["type"]=="remboursement")
    credit = total_dep - total_remb
    lr = len(transactions)+6
    ws.merge_cells(f"A{lr}:E{lr}")
    ws[f"A{lr}"].value = "RESUME"; ws[f"A{lr}"].font = Font(bold=True,color=white)
    ws[f"A{lr}"].fill = PatternFill("solid",fgColor=green); ws[f"A{lr}"].alignment = Alignment(horizontal="center")
    for j,(label,val) in enumerate([("Total depenses",total_dep),("Total remboursements",total_remb),("Credit restant",credit)],lr+1):
        ws.cell(row=j,column=1,value=label).font = Font(bold=True)
        ws.cell(row=j,column=1).fill = PatternFill("solid",fgColor=gray)
        ws.cell(row=j,column=2,value=val).font = Font(bold=True)
        for col in range(1,6): ws.cell(row=j,column=col).border = bd()
    for col,w in zip("ABCDE",[14,20,32,18,18]): ws.column_dimensions[col].width = w
    filename = f"rapport_{client['nom'].replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    return filename

HTML = r"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<title>AgriGest</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Segoe UI,sans-serif;background:#f0f4f0;color:#222}
header{background:#2E7D32;color:#fff;padding:14px 24px;display:flex;align-items:center;gap:12px}
header h1{font-size:20px}
header p{font-size:12px;opacity:.8}
.main{max-width:1050px;margin:20px auto;padding:0 14px}
.stats{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:18px}
.stat{background:#fff;border-radius:8px;padding:14px;border-left:4px solid}
.stat .v{font-size:18px;font-weight:700;margin-top:4px}
.stat .l{font-size:11px;color:#888;text-transform:uppercase}
.s1{border-color:#2E7D32}.s1 .v{color:#2E7D32}
.s2{border-color:#C62828}.s2 .v{color:#C62828}
.s3{border-color:#1565C0}.s3 .v{color:#1565C0}
.s4{border-color:#E65100}.s4 .v{color:#E65100}
.card{background:#fff;border-radius:8px;padding:18px;margin-bottom:16px}
.card h2{font-size:15px;font-weight:600;color:#2E7D32;margin-bottom:14px;padding-bottom:8px;border-bottom:1px solid #eee}
.row{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:10px}
input,select{flex:1;min-width:120px;padding:8px 10px;border:1px solid #ddd;border-radius:6px;font-size:13px}
input:focus,select:focus{outline:none;border-color:#2E7D32}
.btn{padding:8px 16px;border:none;border-radius:6px;font-size:13px;cursor:pointer;font-weight:500}
.bg{background:#2E7D32;color:#fff}.bg:hover{background:#1B5E20}
.br{background:#C62828;color:#fff}.br:hover{background:#8e0000}
.bb{background:#1565C0;color:#fff}.bb:hover{background:#0d47a1}
.bo{background:#E65100;color:#fff}.bo:hover{background:#bf360c}
.sm{padding:4px 9px;font-size:12px}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:#E8F5E9;color:#2E7D32;font-weight:600;padding:9px 10px;text-align:left;border-bottom:2px solid #C8E6C9}
td{padding:8px 10px;border-bottom:1px solid #f0f0f0;vertical-align:middle}
tr:hover td{background:#f9fbe7}
.bdg{display:inline-block;padding:2px 8px;border-radius:99px;font-size:11px;font-weight:600}
.bdr{background:#FFEBEE;color:#C62828}
.bdg2{background:#E8F5E9;color:#2E7D32}
.cn{color:#C62828;font-weight:700}
.cp{color:#2E7D32;font-weight:700}
.lnk{color:#1565C0;cursor:pointer;text-decoration:underline;font-weight:600}
.back{color:#2E7D32;cursor:pointer;font-size:13px;margin-bottom:10px;display:inline-block}
.back:hover{text-decoration:underline}
.empty{text-align:center;color:#aaa;padding:28px;font-size:13px}
#msg{position:fixed;bottom:20px;right:20px;padding:10px 18px;border-radius:8px;font-size:13px;display:none;color:#fff;z-index:99}
.mok{background:#2E7D32}.merr{background:#C62828}
@media(max-width:600px){.stats{grid-template-columns:repeat(2,1fr)}.row{flex-direction:column}}
</style>
</head>
<body>
<header>
  <div style="font-size:26px">&#127807;</div>
  <div><h1>AgriGest</h1><p>Gestion des depenses agricoles par client</p></div>
</header>
<div class="main" id="root">Chargement...</div>
<div id="msg"></div>
<script>
var view='home', curId=null, clients=[], txs=[];

function msg(t,err){
  var d=document.getElementById('msg');
  d.textContent=t; d.className=err?'merr':'mok'; d.style.display='block';
  setTimeout(function(){d.style.display='none'},2800);
}

function fmt(n){
  return Number(n).toLocaleString('fr-FR')+' FCFA';
}

function req(url,method,data,cb){
  var x=new XMLHttpRequest();
  x.open(method,url);
  x.setRequestHeader('Content-Type','application/json');
  x.onload=function(){cb(JSON.parse(x.responseText))};
  x.onerror=function(){msg('Erreur reseau',true)};
  x.send(data?JSON.stringify(data):null);
}

function go(v,id){
  view=v; curId=id||null;
  if(v==='home') req('/api/clients','GET',null,function(d){clients=d;draw()});
  else req('/api/transactions/'+id,'GET',null,function(d){txs=d;draw()});
}

function draw(){
  var r=document.getElementById('root');
  if(view==='home') r.innerHTML=homeHTML();
  else r.innerHTML=detailHTML();
}

function homeHTML(){
  var td=0,tr2=0,tc=0;
  for(var i=0;i<clients.length;i++){td+=clients[i].total_depenses;tr2+=clients[i].total_remboursements;tc+=clients[i].credit_restant;}
  var h='<div class="stats">';
  h+='<div class="stat s1"><div class="l">Clients</div><div class="v">'+clients.length+'</div></div>';
  h+='<div class="stat s2"><div class="l">Total depenses</div><div class="v">'+fmt(td)+'</div></div>';
  h+='<div class="stat s3"><div class="l">Total rembourse</div><div class="v">'+fmt(tr2)+'</div></div>';
  h+='<div class="stat s4"><div class="l">Credit total restant</div><div class="v">'+fmt(tc)+'</div></div>';
  h+='</div>';
  h+='<div class="card"><h2>Ajouter un client</h2>';
  h+='<div class="row">';
  h+='<input id="nom" placeholder="Nom du client *">';
  h+='<input id="tel" placeholder="Telephone">';
  h+='<input id="cul" placeholder="Culture (ex: mais, cacao...)">';
  h+='<button class="btn bg" onclick="addClient()">Ajouter</button>';
  h+='</div></div>';
  h+='<div class="card"><h2>Liste des clients</h2>';
  if(clients.length===0){h+='<div class="empty">Aucun client. Ajoutez votre premier client ci-dessus.</div>';}
  else{
    h+='<table><thead><tr><th>Nom</th><th>Culture</th><th>Telephone</th><th>Depenses</th><th>Rembourse</th><th>Credit restant</th><th>Actions</th></tr></thead><tbody>';
    for(var i=0;i<clients.length;i++){
      var c=clients[i];
      h+='<tr>';
      h+='<td><span class="lnk" onclick="go(\'detail\','+c.id+')">'+c.nom+'</span></td>';
      h+='<td>'+(c.culture||'&mdash;')+'</td>';
      h+='<td>'+(c.telephone||'&mdash;')+'</td>';
      h+='<td>'+fmt(c.total_depenses)+'</td>';
      h+='<td>'+fmt(c.total_remboursements)+'</td>';
      h+='<td class="'+(c.credit_restant>0?'cn':'cp')+'">'+fmt(c.credit_restant)+'</td>';
      h+='<td style="display:flex;gap:5px;flex-wrap:wrap">';
      h+='<button class="btn bb sm" onclick="go(\'detail\','+c.id+')">Detail</button>';
      h+='<button class="btn bo sm" onclick="exporter('+c.id+')">Excel</button>';
      h+='<button class="btn br sm" onclick="supprClient('+c.id+',\''+c.nom+'\')">Suppr.</button>';
      h+='</td></tr>';
    }
    h+='</tbody></table>';
  }
  h+='</div>';
  return h;
}

function detailHTML(){
  var cl=null;
  for(var i=0;i<clients.length;i++) if(clients[i].id===curId){cl=clients[i];break;}
  var td=0,tr2=0;
  for(var i=0;i<txs.length;i++){
    if(txs[i].type==='depense') td+=txs[i].montant;
    else tr2+=txs[i].montant;
  }
  var credit=td-tr2;
  var nom=cl?cl.nom:'Client';
  var h='<span class="back" onclick="go(\'home\')">&#8592; Retour a la liste</span>';
  h+='<div class="stats">';
  h+='<div class="stat s1"><div class="l">Client</div><div class="v" style="font-size:16px">'+nom+'</div></div>';
  h+='<div class="stat s2"><div class="l">Total depenses</div><div class="v">'+fmt(td)+'</div></div>';
  h+='<div class="stat s3"><div class="l">Total rembourse</div><div class="v">'+fmt(tr2)+'</div></div>';
  h+='<div class="stat s4"><div class="l">Credit restant</div><div class="v">'+fmt(credit)+'</div></div>';
  h+='</div>';
  h+='<div class="card"><h2>Nouvelle transaction</h2><div class="row">';
  h+='<select id="ttype"><option value="depense">Depense</option><option value="remboursement">Remboursement</option></select>';
  h+='<input id="tmnt" type="number" placeholder="Montant (FCFA)" min="0">';
  h+='<input id="tdsc" placeholder="Description (ex: engrais, semences...)">';
  h+='<input id="tdat" type="date" value="'+new Date().toISOString().split("T")[0]+'">';
  h+='<button class="btn bg" onclick="addTx()">Enregistrer</button>';
  h+='</div></div>';
  h+='<div class="card">';
  h+='<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;padding-bottom:8px;border-bottom:1px solid #eee">';
  h+='<h2 style="margin:0;border:none;padding:0">Historique des transactions</h2>';
  h+='<button class="btn bo sm" onclick="exporter('+curId+')">Exporter Excel</button>';
  h+='</div>';
  if(txs.length===0){h+='<div class="empty">Aucune transaction enregistree.</div>';}
  else{
    h+='<table><thead><tr><th>Date</th><th>Type</th><th>Description</th><th>Montant</th><th>Action</th></tr></thead><tbody>';
    for(var i=0;i<txs.length;i++){
      var t=txs[i];
      h+='<tr><td>'+t.date+'</td>';
      h+='<td><span class="bdg '+(t.type==='depense'?'bdr':'bdg2')+'">'+(t.type==='depense'?'Depense':'Remboursement')+'</span></td>';
      h+='<td>'+(t.description||'&mdash;')+'</td>';
      h+='<td style="font-weight:600">'+fmt(t.montant)+'</td>';
      h+='<td><button class="btn br sm" onclick="supprTx('+t.id+')">Suppr.</button></td></tr>';
    }
    h+='</tbody></table>';
  }
  h+='</div>';
  return h;
}

function addClient(){
  var nom=document.getElementById('nom').value.trim();
  if(!nom){msg('Le nom est obligatoire',true);return;}
  var tel=document.getElementById('tel').value.trim();
  var cul=document.getElementById('cul').value.trim();
  req('/api/clients','POST',{nom:nom,telephone:tel,culture:cul},function(){
    msg('Client ajoute !'); go('home');
  });
}

function supprClient(id,nom){
  if(!confirm('Supprimer "'+nom+'" et toutes ses transactions ?')) return;
  req('/api/clients/'+id,'DELETE',null,function(){msg('Client supprime');go('home');});
}

function addTx(){
  var type=document.getElementById('ttype').value;
  var mnt=parseFloat(document.getElementById('tmnt').value);
  var dsc=document.getElementById('tdsc').value.trim();
  var dat=document.getElementById('tdat').value;
  if(!mnt||mnt<=0){msg('Montant invalide',true);return;}
  req('/api/transactions','POST',{client_id:curId,type:type,montant:mnt,description:dsc,date:dat},function(){
    msg('Transaction enregistree !');
    req('/api/clients','GET',null,function(d){clients=d;});
    go('detail',curId);
  });
}

function supprTx(id){
  if(!confirm('Supprimer cette transaction ?')) return;
  req('/api/transactions/'+id,'DELETE',null,function(){
    msg('Transaction supprimee');
    req('/api/clients','GET',null,function(d){clients=d;});
    go('detail',curId);
  });
}

function exporter(id){
  msg('Export en cours...');
  req('/api/export/'+id,'GET',null,function(d){
    if(d.file){window.open('/download/'+d.file);msg('Fichier Excel exporte !');}
    else msg('Erreur export',true);
  });
}

go('home');
</script>
</body>
</html>"""

class Handler(BaseHTTPRequestHandler):
    def log_message(self, format, *args): pass

    def send_json(self, data, code=200):
        body = json.dumps(data, ensure_ascii=False).encode()
        self.send_response(code)
        self.send_header("Content-Type","application/json; charset=utf-8")
        self.send_header("Content-Length",len(body))
        self.end_headers(); self.wfile.write(body)

    def do_GET(self):
        from urllib.parse import urlparse
        path = urlparse(self.path).path
        if path in ("/","/index.html"):
            body = HTML.encode()
            self.send_response(200)
            self.send_header("Content-Type","text/html; charset=utf-8")
            self.send_header("Content-Length",len(body))
            self.end_headers(); self.wfile.write(body)
        elif path=="/api/clients":
            self.send_json(get_clients())
        elif path.startswith("/api/transactions/"):
            self.send_json(get_transactions(int(path.split("/")[-1])))
        elif path.startswith("/api/export/"):
            f = export_excel(int(path.split("/")[-1]))
            self.send_json({"file":f} if f else {"error":"not found"})
        elif path.startswith("/download/"):
            fname = path[len("/download/"):]
            if os.path.exists(fname):
                with open(fname,"rb") as f: data=f.read()
                self.send_response(200)
                self.send_header("Content-Type","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition",f'attachment; filename="{fname}"')
                self.send_header("Content-Length",len(data))
                self.end_headers(); self.wfile.write(data)
            else: self.send_json({"error":"not found"},404)
        else: self.send_json({"error":"not found"},404)

    def do_POST(self):
        length = int(self.headers.get("Content-Length",0))
        body = json.loads(self.rfile.read(length))
        if self.path=="/api/clients":
            add_client(body["nom"],body.get("telephone",""),body.get("culture",""))
            self.send_json({"ok":True})
        elif self.path=="/api/transactions":
            add_transaction(body["client_id"],body["type"],body["montant"],
                body.get("description",""),body.get("date",str(datetime.now().date())))
            self.send_json({"ok":True})
        else: self.send_json({"error":"not found"},404)

    def do_DELETE(self):
        if self.path.startswith("/api/clients/"):
            delete_client(int(self.path.split("/")[-1])); self.send_json({"ok":True})
        elif self.path.startswith("/api/transactions/"):
            delete_transaction(int(self.path.split("/")[-1])); self.send_json({"ok":True})
        else: self.send_json({"error":"not found"},404)

if __name__=="__main__":
    init_db()
    PORT=5000
    import socket, qrcode

    # Trouver l'adresse IP locale automatiquement
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    s.connect(("8.8.8.8", 80))
    IP = s.getsockname()[0]
    s.close()

    URL = f"http://{IP}:{PORT}"

    # Générer et afficher le QR code dans le terminal
    qr = qrcode.QRCode(border=2)
    qr.add_data(URL)
    qr.make(fit=True)
    qr.print_ascii(invert=True)

    print(f"""
+------------------------------------------+
|        AgriGest demarre !                |
|  PC      : http://localhost:{PORT}          |
|  Telephone : {URL}  |
|  Scanne le QR code avec ton telephone    |
|  Ctrl+C pour arreter                     |
+------------------------------------------+""")

    server=HTTPServer(("0.0.0.0",PORT),Handler)
    threading.Timer(1.5,lambda: webbrowser.open("http://localhost:"+str(PORT))).start()
    try: server.serve_forever()
    except KeyboardInterrupt: print("\nAgriGest arrete.")
